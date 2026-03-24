import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO

st.set_page_config(page_title="财务专用-保留公式版", layout="wide")
st.title("📊 财务报表对应填充 (保留公式版)")

source_file = st.file_uploader("1. 上传【数据源表】(提取数据用)", type=['xlsx'])
target_file = st.file_uploader("2. 上传【目标模板表】(带公式的表)", type=['xlsx'])

if source_file and target_file:
    # 读取源表数据
    df_source = pd.read_excel(source_file)
    
    # 使用 openpyxl 加载目标表，保留公式
    wb_target = openpyxl.load_workbook(target_file, data_only=False)
    sheet = wb_target.active
    
    # 获取表头
    target_headers = [cell.value for cell in sheet[1]]
    
    st.write("### 配置匹配规则")
    col1, col2 = st.columns(2)
    with col1:
        s_key = st.selectbox("源表关联列（如：编号）", df_source.columns)
    with col2:
        t_key = st.selectbox("目标表关联列（必须在第一行）", target_headers)
    
    # 找到关联列在目标表是第几列 (从1开始)
    t_key_idx = target_headers.index(t_key) + 1
    
    # 选择要填充的列
    cols_to_fill = st.multiselect("选择要从源表同步到目标表的列", 
                                  [c for c in df_source.columns if c != s_key])

    if st.button("开始保留公式填充"):
        # 将源表转为字典，方便快速查询
        source_dict = df_source.set_index(s_key).to_dict('index')
        
        # 遍历目标表每一行 (从第2行开始)
        fill_count = 0
        for row_idx in range(2, sheet.max_row + 1):
            key_value = sheet.cell(row=row_idx, column=t_key_idx).value
            
            if key_value in source_dict:
                # 找到匹配的数据了
                match_data = source_dict[key_value]
                for col_name in cols_to_fill:
                    # 找到目标表对应的列号
                    if col_name in target_headers:
                        t_col_idx = target_headers.index(col_name) + 1
                        # 填充值
                        sheet.cell(row=row_idx, column=t_col_idx).value = match_data[col_name]
                        fill_count += 1

        st.success(f"填充完成！共填入 {fill_count} 条数据，原有公式已保留。")

        # 导出文件
        output = BytesIO()
        wb_target.save(output)
        st.download_button("📩 下载保留公式后的表格", output.getvalue(), "财务填充结果.xlsx")
